#!/usr/bin/env python3
"""
ドンキ担当ベンダー 業務引き継ぎ表テンプレート生成スクリプト
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

def create_handover_template():
    wb = Workbook()

    # スタイル定義
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    wrap_alignment = Alignment(wrap_text=True, vertical='top')

    def setup_sheet(ws, headers, col_widths):
        """シートの基本設定"""
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')

        for col_idx, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # 2行目以降の書式設定
        for row in range(2, 100):
            for col in range(1, len(headers) + 1):
                ws.cell(row=row, column=col).border = thin_border
                ws.cell(row=row, column=col).alignment = wrap_alignment

    # ========================================
    # シート1: 引き継ぎサマリー
    # ========================================
    ws_summary = wb.active
    ws_summary.title = "引き継ぎサマリー"

    summary_items = [
        ("作成日", datetime.now().strftime("%Y/%m/%d")),
        ("引き継ぎ元担当者", ""),
        ("引き継ぎ先担当者", ""),
        ("担当店舗数", ""),
        ("進行中タスク数", ""),
        ("未解決課題数", ""),
        ("", ""),
        ("全体サマリー", ""),
    ]

    for row_idx, (label, value) in enumerate(summary_items, 1):
        ws_summary.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
        ws_summary.cell(row=row_idx, column=1).fill = header_fill
        ws_summary.cell(row=row_idx, column=1).border = thin_border
        ws_summary.cell(row=row_idx, column=2, value=value)
        ws_summary.cell(row=row_idx, column=2).border = thin_border
        ws_summary.cell(row=row_idx, column=2).alignment = wrap_alignment

    ws_summary.column_dimensions['A'].width = 20
    ws_summary.column_dimensions['B'].width = 60
    ws_summary.row_dimensions[8].height = 100  # 全体サマリー用

    # ========================================
    # シート2: 店舗一覧
    # ========================================
    ws_stores = wb.create_sheet("店舗一覧")
    store_headers = ["店舗コード", "店舗名", "エリア", "店舗タイプ", "棚割りパターン",
                     "最終訪問日", "次回予定", "特記事項", "優先度"]
    store_widths = [12, 25, 10, 15, 15, 12, 12, 30, 8]
    setup_sheet(ws_stores, store_headers, store_widths)

    # ========================================
    # シート3: 進行中タスク
    # ========================================
    ws_tasks = wb.create_sheet("進行中タスク")
    task_headers = ["No.", "店舗名", "タスク種別", "タスク内容", "優先度",
                    "ステータス", "期限", "担当者", "関連者", "備考"]
    task_widths = [5, 25, 15, 30, 8, 10, 12, 12, 20, 25]
    setup_sheet(ws_tasks, task_headers, task_widths)

    # ========================================
    # シート4: 課題・問題一覧
    # ========================================
    ws_issues = wb.create_sheet("課題・問題一覧")
    issue_headers = ["No.", "店舗名", "課題カテゴリ", "課題内容", "発生日",
                     "影響度", "対応状況", "暫定対処", "恒久対策", "担当者", "備考"]
    issue_widths = [5, 25, 12, 30, 12, 8, 10, 25, 25, 12, 20]
    setup_sheet(ws_issues, issue_headers, issue_widths)

    # ========================================
    # シート5: 棚割り状況
    # ========================================
    ws_shelf = wb.create_sheet("棚割り状況")
    shelf_headers = ["店舗名", "現在の棚割り", "適用開始日", "次回更新予定",
                     "変更頻度", "季節対応", "特殊対応", "備考"]
    shelf_widths = [25, 20, 12, 12, 12, 10, 25, 25]
    setup_sheet(ws_shelf, shelf_headers, shelf_widths)

    # ========================================
    # シート6: 連絡先一覧
    # ========================================
    ws_contacts = wb.create_sheet("連絡先一覧")
    contact_headers = ["所属", "担当者名", "役職", "電話番号", "メール", "備考"]
    contact_widths = [15, 15, 15, 15, 25, 25]
    setup_sheet(ws_contacts, contact_headers, contact_widths)

    # ========================================
    # ファイル保存
    # ========================================
    filename = "業務引き継ぎ表.xlsx"
    wb.save(filename)
    print(f"✅ {filename} を作成しました")
    return filename

if __name__ == "__main__":
    create_handover_template()
